#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
02_clean.py — Nettoyage par source de données
Exécution : python scripts/02_clean.py

Chaque fonction de nettoyage est indépendante : si une source est absente,
un avertissement est affiché et l'exécution continue normalement.

Sorties dans data/processed/ :
  surendettement.csv   — BdF API WebStat 2019–2024 (mensuel agrégé annuel)
  filosofi.csv         — FiLoSoFi 2021 (revenu, pauvreté, interdécile)
  filosofi_gini.csv    — FiLoSoFi SUPRA 2019 (Gini)
  chomage.csv          — Chômage localisé INSEE
  rp_pop_ref.csv       — RP 2022 populations de référence (population municipale par département)
  rp_infracommunal.csv — RP 2022 bases infracommunales (locataires, HLM, surpeuplement, familles mono…)
  minimas_sociaux.csv  — RSA, prime d'activité, ASS/ASPA
"""

import pathlib
import re
import unicodedata
import warnings

import pandas as pd

ROOT      = pathlib.Path(__file__).parent.parent
RAW       = ROOT / "data" / "raw"
PROCESSED = ROOT / "data" / "processed"
PROCESSED.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# Utilitaires partagés
# ---------------------------------------------------------------------------

def load_dep_ref() -> pd.DataFrame:
    """Charge la table de référence des 96 départements."""
    return pd.read_csv(PROCESSED / "dep_ref.csv", dtype=str)


def normalize_name(name: str) -> str:
    """Normalise un nom de département : minuscules, sans accents, tirets → espaces."""
    if not isinstance(name, str):
        return ""
    # décomposer les caractères accentués
    nfkd = unicodedata.normalize("NFKD", name.lower())
    ascii_str = "".join(c for c in nfkd if not unicodedata.combining(c))
    # unifier séparateurs
    ascii_str = re.sub(r"[-_''`]", " ", ascii_str)
    ascii_str = re.sub(r"\s+", " ", ascii_str).strip()
    return ascii_str


# ---------------------------------------------------------------------------
# T011 — BdF : API WebStat CSV → surendettement.csv
# ---------------------------------------------------------------------------

def clean_surendettement():
    """Agrège les données mensuelles BdF (API WebStat) en totaux annuels par département.

    Lit data/raw/bdf/surendettement_api.csv (délimiteur ;), extrait l'année
    depuis time_period (YYYY-MM), somme les obs_value mensuelles par
    (dep_code, annee), et filtre les années complètes (2019–2024).
    Produit data/processed/surendettement.csv.
    """
    print("\n=== Nettoyage BdF — surendettement ===")

    api_csv = RAW / "bdf" / "surendettement_api.csv"
    if not api_csv.exists():
        print(f"  ⚠️  {api_csv.name} absent — ignorer bloc BdF")
        pd.DataFrame(
            columns=["dep_code", "dep_nom", "annee", "suren_depot_nb",
                     "source_url", "source_millesime"]
        ).to_csv(PROCESSED / "surendettement.csv", index=False)
        return

    dep_ref = load_dep_ref()
    print(f"  ↳ Chargement {api_csv.name} …")
    df = pd.read_csv(api_csv, sep=";", encoding="utf-8-sig")

    # series_key "IFI.M.D01.SUREN.DEPOT" → dep_code "01", "2A", etc.
    df["dep_code"] = df["series_key"].str.split(".").str[2].str[1:]

    # time_period "2019-01" → annee 2019
    df["annee"] = df["time_period"].str[:4].astype(int)

    # Garder uniquement les années complètes 2019–2024
    ANNEES_COMPLETES = list(range(2019, 2025))
    df = df[df["annee"].isin(ANNEES_COMPLETES)]

    # Agréger : somme mensuelle → total annuel par département
    df_annual = (
        df.groupby(["dep_code", "annee"], as_index=False)["obs_value"]
        .sum()
        .rename(columns={"obs_value": "suren_depot_nb"})
    )
    df_annual["suren_depot_nb"] = df_annual["suren_depot_nb"].round().astype(int)

    # Joindre les noms de département
    df_annual = df_annual.merge(dep_ref[["dep_code", "dep_nom"]], on="dep_code", how="left")

    # Métadonnées source
    df_annual["source_url"] = df_annual["dep_code"].apply(
        lambda c: f"https://webstat.banque-france.fr/fr/catalogue/ifi/IFI.M.D{c}.SUREN.DEPOT"
    )
    df_annual["source_millesime"] = df_annual["annee"].astype(str)

    cols = ["dep_code", "dep_nom", "annee", "suren_depot_nb", "source_url", "source_millesime"]
    df_annual = df_annual[cols].sort_values(["annee", "dep_code"])

    # Contrôle qualité
    nb_dep_par_annee = df_annual.groupby("annee")["dep_code"].count()
    for annee, nb in nb_dep_par_annee.items():
        flag = "✓" if nb == 96 else "⚠️ "
        print(f"  {flag} {annee} : {nb}/96 départements")

    df_annual.to_csv(PROCESSED / "surendettement.csv", index=False)
    print(f"  ✓ surendettement.csv — {len(df_annual)} lignes")


# ---------------------------------------------------------------------------
# T012 — FiLoSoFi : revenu médian, pauvreté, interdécile, Gini
# ---------------------------------------------------------------------------

def clean_filosofi():
    """Nettoie les données FiLoSoFi 2023 (format long ODS/SDMX).

    Source : DS_FILOSOFI_CC_2023_data.csv (long format, `;` délimiteur)
    Mesures extraites au niveau DEP (GEO_OBJECT == 'DEP') :
      MED_SL       → revenu_median_uc   (niveau de vie médian, €)
      PR_MD60      → taux_pauvrete      (taux de pauvreté seuil 60 %, %)
      IR_D9_D1_SL  → interdecile_d9d1   (rapport D9/D1)
      GI_SL        → gini               (indice de Gini) → filosofi_gini.csv

    Pas de colonne `annee` → jointure sur dep_code seul dans 03_merge.py
    (les valeurs 2023 servent de référence structurelle pour toutes les années).
    """
    print("\n=== Nettoyage FiLoSoFi ===")

    filosofi_dir = RAW / "filosofi"
    empty_filo = pd.DataFrame(
        columns=["dep_code", "revenu_median_uc", "taux_pauvrete",
                 "interdecile_d9d1", "source_url", "source_millesime"]
    )
    empty_gini = pd.DataFrame(
        columns=["dep_code", "gini", "source_url", "source_millesime"]
    )

    data_file = filosofi_dir / "DS_FILOSOFI_CC_2023_data.csv"
    if not data_file.exists():
        print(f"  ⚠️  Fichier FiLoSoFi 2023 non trouvé dans {filosofi_dir}")
        empty_filo.to_csv(PROCESSED / "filosofi.csv", index=False)
        empty_gini.to_csv(PROCESSED / "filosofi_gini.csv", index=False)
        return

    print(f"  ↳ Chargement {data_file.name} …")
    try:
        df = pd.read_csv(data_file, sep=";", dtype=str, low_memory=False)
    except Exception as exc:
        print(f"  ⚠️  Erreur de lecture FiLoSoFi 2023 : {exc}")
        empty_filo.to_csv(PROCESSED / "filosofi.csv", index=False)
        empty_gini.to_csv(PROCESSED / "filosofi_gini.csv", index=False)
        return

    # Filtrer les lignes de niveau département et exclure les DOM (971-976)
    df_dep = df[df["GEO_OBJECT"] == "DEP"].copy()
    df_dep = df_dep[df_dep["GEO"].str.match(r"^([0-9]{2}|2[AB])$", na=False)]

    # Pivoter : une ligne par département
    MEASURE_MAP = {
        "MED_SL":      "revenu_median_uc",
        "PR_MD60":     "taux_pauvrete",
        "IR_D9_D1_SL": "interdecile_d9d1",
        "GI_SL":       "gini",
    }
    df_dep = df_dep[df_dep["FILOSOFI_MEASURE"].isin(MEASURE_MAP)]
    df_dep["variable"] = df_dep["FILOSOFI_MEASURE"].map(MEASURE_MAP)
    df_dep["OBS_VALUE"] = pd.to_numeric(df_dep["OBS_VALUE"], errors="coerce")
    df_pivot = df_dep.pivot_table(
        index="GEO", columns="variable", values="OBS_VALUE", aggfunc="first"
    ).reset_index().rename(columns={"GEO": "dep_code"})

    SOURCE_URL = "https://www.insee.fr/fr/statistiques/8984752"
    df_pivot["source_url"]       = SOURCE_URL
    df_pivot["source_millesime"] = "2023"

    # ── filosofi.csv (revenu, pauvreté, interdécile) ─────────────────────────
    filo_cols = ["dep_code"] + [
        c for c in ["revenu_median_uc", "taux_pauvrete", "interdecile_d9d1"]
        if c in df_pivot.columns
    ] + ["source_url", "source_millesime"]
    filo_out = df_pivot[filo_cols].copy()
    filo_out.to_csv(PROCESSED / "filosofi.csv", index=False)
    print(f"  ✓ filosofi.csv — {len(filo_out)} lignes, colonnes : {filo_out.columns.tolist()}")

    # ── filosofi_gini.csv ─────────────────────────────────────────────────────
    if "gini" in df_pivot.columns:
        gini_out = df_pivot[["dep_code", "gini", "source_url", "source_millesime"]].copy()
        gini_out = gini_out[gini_out["gini"].notna()]
        gini_out.to_csv(PROCESSED / "filosofi_gini.csv", index=False)
        print(f"  ✓ filosofi_gini.csv — {len(gini_out)} lignes")
    else:
        empty_gini.to_csv(PROCESSED / "filosofi_gini.csv", index=False)
        print("  ⚠️  Aucun fichier FiLoSoFi SUPRA 2019 trouvé — filosofi_gini.csv vide")


# ---------------------------------------------------------------------------
# T013 — Chômage localisé INSEE → chomage.csv
# ---------------------------------------------------------------------------

def clean_chomage():
    """Nettoie les données de chômage localisé INSEE."""
    print("\n=== Nettoyage Chômage localisé ===")

    chomage_dir = RAW / "chomage"
    # Chercher un fichier Excel TCRD
    excel_files = (
        list(chomage_dir.glob("**/*.xlsx"))
        + list(chomage_dir.glob("**/*.xls"))
        + list(chomage_dir.glob("**/*.csv"))
    )
    # Filtrer les fichiers pertinents (TCRD ou taux-chomage)
    excel_files = [
        f for f in excel_files
        if any(kw in f.stem.upper()
               for kw in ["TCRD", "CHOM", "TAUX", "LOCALIS"])
    ] or excel_files  # si aucun filtre, prendre tous

    if not excel_files:
        print(f"  ⚠️  Aucun fichier chômage dans {chomage_dir} — chomage.csv vide")
        print("  💡 Télécharger depuis : https://www.insee.fr/fr/statistiques/2012804")
        pd.DataFrame(
            columns=["dep_code", "chomage_taux",
                     "source_url", "source_millesime"]
        ).to_csv(PROCESSED / "chomage.csv", index=False)
        return

    filepath = excel_files[0]
    print(f"  ↳ Chargement {filepath.name} …")

    try:
        if filepath.suffix in (".xlsx", ".xls"):
            df_raw = pd.read_excel(filepath, header=None, dtype=str)
        else:
            df_raw = pd.read_csv(filepath, dtype=str, sep=None, engine="python")

        # ── Détecter la structure du fichier ──────────────────────────────
        # Les fichiers TCRD INSEE ont généralement :
        # - Lignes d'en-tête (codes trimestres ou années)
        # - Colonne 0 : code département (01..95, 2A, 2B)
        # - Colonnes suivantes : taux de chômage par trimestre

        # Trouver la première ligne avec des codes département
        dep_mask_regex = r"^([0-9]{2}|2[AB])$"
        header_row = None
        for i, row in df_raw.iterrows():
            if row.iloc[0] in ["01", "02", "1", "2"] or re.match(
                dep_mask_regex, str(row.iloc[0])
            ):
                header_row = i
                break

        if header_row is None:
            print("  ⚠️  Structure du fichier TCRD non reconnue")
            pd.DataFrame(
                columns=["dep_code", "chomage_taux",
                         "source_url", "source_millesime"]
            ).to_csv(PROCESSED / "chomage.csv", index=False)
            return

        # En-têtes : ligne précédente
        col_headers = df_raw.iloc[header_row - 1].tolist() if header_row > 0 else []
        df = df_raw.iloc[header_row:].copy()
        df.columns = range(len(df.columns))

        # Filtrer lignes département (96 métropolitains, exclure 971-976)
        mask = df[0].apply(
            lambda x: bool(re.match(dep_mask_regex, str(x).strip()))
        )
        df = df[mask].copy()
        df = df.rename(columns={0: "dep_code"})
        df["dep_code"] = df["dep_code"].str.zfill(2)

        # Convertir colonnes numériques en taux
        numeric_cols = df.columns[1:]
        records: list[dict] = []

        # Identifier la colonne Q4 2024 si présente, sinon prendre la dernière colonne
        ref_col_idx = None
        if header_row and header_row > 0:
            headers_row = df_raw.iloc[header_row - 1].tolist()
            for ci, h in enumerate(headers_row):
                if h and "2024" in str(h):
                    ref_col_idx = ci  # index de colonne dans df (0-basé, identique au raw)
                    break

        for _, row in df.iterrows():
            dep_code = row["dep_code"]
            # Priorité : colonne de référence Q4 2024 ; sinon moyenne de toutes
            if ref_col_idx is not None and ref_col_idx in numeric_cols:
                try:
                    val = float(str(row[ref_col_idx]).replace(",", ".").strip())
                    chomage_taux = val if 0 < val < 30 else None
                except (ValueError, TypeError):
                    chomage_taux = None
            else:
                chomage_taux = None

            if chomage_taux is None:
                # Repli : moyenne de toutes les colonnes numériques disponibles
                values: list[float] = []
                for col in numeric_cols:
                    try:
                        val = float(str(row[col]).replace(",", ".").strip())
                        if 0 < val < 30:
                            values.append(val)
                    except (ValueError, TypeError):
                        pass
                chomage_taux = round(sum(values) / len(values), 2) if values else None

            if chomage_taux is not None:
                records.append({
                    "dep_code":         dep_code,
                    # Pas d'annee → jointure sur dep_code seul dans 03_merge.py
                    # (taux de référence Q4 2024 diffusé sur toutes les années)
                    "chomage_taux":     round(chomage_taux, 2),
                    "source_url":       "https://www.insee.fr/fr/statistiques/2012804",
                    "source_millesime": "2024",
                })

        if records:
            result = pd.DataFrame(records)
            result.to_csv(PROCESSED / "chomage.csv", index=False)
            print(f"  ✓ chomage.csv — {len(result)} lignes")
        else:
            print("  ⚠️  Aucune donnée extraite du fichier TCRD")
            pd.DataFrame(
                columns=["dep_code", "chomage_taux",
                         "source_url", "source_millesime"]
            ).to_csv(PROCESSED / "chomage.csv", index=False)

    except Exception as exc:
        print(f"  ✗ Erreur nettoyage chômage : {exc}")
        pd.DataFrame(
            columns=["dep_code", "chomage_taux",
                     "source_url", "source_millesime"]
        ).to_csv(PROCESSED / "chomage.csv", index=False)


# ---------------------------------------------------------------------------
# T014 — RP 2022 : populations de référence + bases infracommunales
# Sources :
#   - Pop. de référence  : https://www.insee.fr/fr/statistiques/8290607?sommaire=8290669
#   - Logement (IRIS)    : https://www.insee.fr/fr/statistiques/8647012
#   - Ménages (IRIS)     : https://www.insee.fr/fr/statistiques/8647008
#   - Population (IRIS)  : https://www.insee.fr/fr/statistiques/8647014
# ---------------------------------------------------------------------------

def _read_iris_csv(zip_path: pathlib.Path, usecols: list[str]) -> pd.DataFrame:
    """Lit le premier CSV dans un ZIP IRIS, en ne chargeant que les colonnes utiles."""
    import zipfile
    with zipfile.ZipFile(zip_path) as z:
        csvs = [n for n in z.namelist() if n.lower().endswith(".csv")
                and not n.lower().startswith("meta")]
        if not csvs:
            raise FileNotFoundError(f"Aucun CSV dans {zip_path.name}")
        with z.open(csvs[0]) as f:
            df = pd.read_csv(
                f, sep=";", dtype=str, low_memory=False,
                usecols=lambda c: c in (["COM"] + usecols),
            )
    return df


def clean_rp():
    """Extrait les indicateurs RP 2022 par département (pop. référence + infracommunales)."""
    print("\n=== Nettoyage RP 2022 (populations de référence) ===")

    rp_dir = RAW / "rp"
    dep_ref = load_dep_ref()

    # ── 1. Populations de référence (agrégat départemental) ─────────────────
    dep_csv = rp_dir / "donnees_departements.csv"

    if not dep_csv.exists():
        import zipfile
        zip_path = list(rp_dir.glob("ensemble.zip"))
        if zip_path:
            with zipfile.ZipFile(zip_path[0]) as z:
                z.extract("donnees_departements.csv", rp_dir)
            dep_csv = rp_dir / "donnees_departements.csv"

    if not dep_csv.exists():
        print(f"  ⚠️  donnees_departements.csv non trouvé dans {rp_dir}")
        pd.DataFrame(columns=["dep_code", "population_mun"]).to_csv(
            PROCESSED / "rp_pop_ref.csv", index=False
        )
        return

    df_ref = pd.read_csv(dep_csv, sep=";", dtype=str)
    df_ref = df_ref.rename(columns={"DEP": "dep_code", "PMUN": "population_mun"})
    df_ref["population_mun"] = pd.to_numeric(df_ref["population_mun"], errors="coerce")

    metro = dep_ref["dep_code"].tolist() if not dep_ref.empty else []
    if metro:
        df_ref = df_ref[df_ref["dep_code"].isin(metro)].copy()
    else:
        df_ref = df_ref[df_ref["dep_code"].str.match(r"^([0-9]{2}|2[AB])$", na=False)].copy()

    df_ref = df_ref[["dep_code", "population_mun"]].copy()
    df_ref["source_url"] = "https://www.insee.fr/fr/statistiques/8290607?sommaire=8290669"
    df_ref["source_millesime"] = "2022"

    df_ref.to_csv(PROCESSED / "rp_pop_ref.csv", index=False)
    print(f"  ✓ rp_pop_ref.csv — {len(df_ref)} départements")

    # ── 2. Bases infracommunales IRIS ────────────────────────────────────────
    print("\n=== Nettoyage RP 2022 (bases infracommunales) ===")

    def agg_dep(zip_path: pathlib.Path, usecols: list[str]) -> pd.DataFrame | None:
        """Agrège les colonnes numériques par département depuis un CSV IRIS zippé."""
        if not zip_path.exists():
            print(f"  ⚠️  {zip_path.name} non trouvé — ignoré")
            return None
        print(f"  ↳ Lecture {zip_path.name} …")
        df = _read_iris_csv(zip_path, usecols)
        if "COM" not in df.columns:
            print(f"  ⚠️  Colonne COM absente dans {zip_path.name}")
            return None
        # Extraire le code département depuis le code commune INSEE (2 premiers chars)
        # Format communes : "01001" → "01", "2A001" → "2A", "2B003" → "2B"
        df["dep_code"] = df["COM"].str[:2]
        for c in usecols:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce")
        df = df.groupby("dep_code", as_index=False)[
            [c for c in usecols if c in df.columns]
        ].sum()
        if metro:
            df = df[df["dep_code"].isin(metro)].copy()
        return df

    # Base logement → locataires, HLM, surpeuplement
    log_cols = ["P22_RP", "P22_RP_LOC", "P22_RP_LOCHLMV",
                "C22_RP_SUROCC_MOD", "C22_RP_SUROCC_ACC"]
    df_log = agg_dep(rp_dir / "base-ic-logement-2022_csv.zip", log_cols)

    # Base ménages/familles → monoparentales, ménages 1 pers.
    men_cols = ["C22_MEN", "C22_MENPSEUL", "C22_MENFAMMONO"]
    df_men = agg_dep(rp_dir / "base-ic-couples-familles-menages-2022_csv.zip", men_cols)

    # Base population → tranches d'âge 25-54 et 65+
    pop_cols = ["P22_POP", "P22_POP2539", "P22_POP4054", "P22_POP65P"]
    df_pop = agg_dep(rp_dir / "base-ic-evol-struct-pop-2022_csv.zip", pop_cols)

    if df_log is None and df_men is None and df_pop is None:
        print("  ⚠️  Aucune base infracommunale disponible — rp_infracommunal.csv vide")
        pd.DataFrame(columns=["dep_code"]).to_csv(
            PROCESSED / "rp_infracommunal.csv", index=False
        )
        return

    # Assemblage
    infra = df_ref[["dep_code"]].copy()
    if df_log is not None:
        infra = infra.merge(df_log, on="dep_code", how="left")
    if df_men is not None:
        infra = infra.merge(df_men, on="dep_code", how="left")
    if df_pop is not None:
        infra = infra.merge(df_pop, on="dep_code", how="left")

    # Calcul des ratios (en %)
    if df_log is not None:
        infra["part_locataires"]     = (infra["P22_RP_LOC"]    / infra["P22_RP"] * 100).round(2)
        infra["part_hlm"]            = (infra["P22_RP_LOCHLMV"] / infra["P22_RP"] * 100).round(2)
        surocc = infra.get("C22_RP_SUROCC_MOD", 0) + infra.get("C22_RP_SUROCC_ACC", 0)
        infra["taux_surpeuplement"]  = (surocc / infra["P22_RP"] * 100).round(2)

    if df_men is not None:
        infra["part_familles_mono"]  = (infra["C22_MENFAMMONO"] / infra["C22_MEN"] * 100).round(2)
        infra["part_menages_1pers"]  = (infra["C22_MENPSEUL"]   / infra["C22_MEN"] * 100).round(2)

    if df_pop is not None:
        pop2554 = infra["P22_POP2539"] + infra["P22_POP4054"]
        infra["part_25_54"]          = (pop2554           / infra["P22_POP"] * 100).round(2)
        infra["part_65plus"]         = (infra["P22_POP65P"] / infra["P22_POP"] * 100).round(2)

    # Garder uniquement les colonnes finales
    final_cols = ["dep_code", "part_locataires", "part_hlm", "taux_surpeuplement",
                  "part_familles_mono", "part_menages_1pers", "part_25_54", "part_65plus"]
    out_cols = [c for c in final_cols if c in infra.columns]
    infra[out_cols].to_csv(PROCESSED / "rp_infracommunal.csv", index=False)

    n_ok = infra["part_locataires"].notna().sum() if "part_locataires" in infra.columns else 0
    print(f"  ✓ rp_infracommunal.csv — {len(infra)} départements, {n_ok} valides pour part_locataires")


# ---------------------------------------------------------------------------
# T015 — Minimas sociaux → minimas_sociaux.csv
# DREES : suivi mensuel des prestations de solidarité (RSA, ASS, prime d'activité)
# Source : https://data.drees.solidarites-sante.gouv.fr/explore/dataset/
#          donnees-mensuelles-sur-les-prestations-de-solidarite/
# ---------------------------------------------------------------------------

def clean_minimas():
    """Parse le XLSX DREES des prestations de solidarité.

    Extrait les effectifs de décembre (snapshot fin d'année) pour 2019-2024
    aux tableaux :
      - Tableau 3 : RSA
      - Tableau 5 : ASS
      - Tableau 6 : prime d'activité
    Normalise par la population municipale (RP 2022) → taux pour 1 000 hab.
    """
    print("\n=== Nettoyage Minimas sociaux (DREES) ===")

    minimas_dir = RAW / "minimas"
    dep_ref = load_dep_ref()

    xlsx_files = list(minimas_dir.glob("*.xlsx")) + list(minimas_dir.glob("*.xls"))
    if not xlsx_files:
        print(f"  ⚠️  Aucun fichier XLSX dans {minimas_dir} — placeholder NaN")
        _write_minimas_placeholder(dep_ref)
        return

    filepath = next(
        (f for f in xlsx_files if "prestations" in f.name.lower() or "drees" in f.name.lower()),
        xlsx_files[0],
    )
    print(f"  ↳ Chargement {filepath.name} …")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as exc:
        print(f"  ✗ Impossible d'ouvrir {filepath.name} : {exc}")
        _write_minimas_placeholder(dep_ref)
        return

    metro_codes = set(dep_ref["dep_code"].tolist()) if not dep_ref.empty else set()

    def parse_tableau(sheet_name: str, col_label: str) -> pd.DataFrame:
        """Parse un tableau DREES dept × mois → snapshot décembre par année."""
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  Feuille {sheet_name} absente")
            return pd.DataFrame()
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 8:
            return pd.DataFrame()

        num_row = all_rows[4]  # Row 5 : numéros département (1, 2, 2A, 2B…)

        # Construire col_idx → dep_code (metropolitain uniquement)
        col_to_dep: dict[int, str] = {}
        for ci, num in enumerate(num_row):
            if num is None:
                continue
            ns = str(num).strip()
            if ns in ("2A", "2B"):
                dep_code = ns
            elif ns.isdigit() and 1 <= int(ns) <= 95:
                dep_code = str(int(ns)).zfill(2)
            else:
                continue
            if not metro_codes or dep_code in metro_codes:
                col_to_dep[ci] = dep_code

        records = []
        for row in all_rows[7:]:
            dt = row[0]
            if dt is None:
                continue
            try:
                if hasattr(dt, "year"):
                    year, month = dt.year, dt.month
                else:
                    parsed = pd.to_datetime(str(dt))
                    year, month = parsed.year, parsed.month
            except Exception:
                continue
            # Snapshot fin d'année (décembre)
            if month != 12 or year < 2019 or year > 2024:
                continue
            for ci, dep_code in col_to_dep.items():
                raw = row[ci]
                if raw is None:
                    continue
                try:
                    val = float(str(raw).replace("*", "").strip())
                    records.append({"annee": year, "dep_code": dep_code, col_label: val})
                except (ValueError, AttributeError):
                    pass
        return pd.DataFrame(records)

    rsa = parse_tableau("Tableau 3", "rsa_nb")
    ass = parse_tableau("Tableau 5", "ass_nb")
    pa  = parse_tableau("Tableau 6", "prime_activite_nb")

    if rsa.empty and ass.empty and pa.empty:
        print("  ⚠️  Aucun tableau parsé — placeholder NaN")
        _write_minimas_placeholder(dep_ref)
        return

    # Fusionner les 3 tableaux
    df = rsa.merge(ass, on=["annee", "dep_code"], how="outer") \
            .merge(pa,  on=["annee", "dep_code"], how="outer")

    # Normaliser par population municipale (RP 2022) → taux pour 1 000 hab.
    pop_ref_path = PROCESSED / "rp_pop_ref.csv"
    if pop_ref_path.exists():
        pop = pd.read_csv(pop_ref_path, dtype={"dep_code": str})
        if "population_mun" in pop.columns:
            df = df.merge(pop[["dep_code", "population_mun"]], on="dep_code", how="left")
            for nb_col, rate_col in [
                ("rsa_nb",             "rsa_taux"),
                ("ass_nb",             "ass_aspa_taux"),
                ("prime_activite_nb",  "prime_activite_taux"),
            ]:
                if nb_col in df.columns:
                    df[rate_col] = (
                        df[nb_col] / df["population_mun"].replace(0, float("nan")) * 1000
                    ).round(3)
            df = df.drop(columns=["population_mun"])
    else:
        print("  ⚠️  rp_pop_ref.csv absent — colonnes _taux non calculées")
        for rate_col in ("rsa_taux", "ass_aspa_taux", "prime_activite_taux"):
            df[rate_col] = float("nan")

    df["source_url"] = (
        "https://data.drees.solidarites-sante.gouv.fr/explore/dataset/"
        "donnees-mensuelles-sur-les-prestations-de-solidarite/"
    )
    df["source_millesime"] = "2019-2024"

    keep = ["dep_code", "annee",
            "rsa_taux", "prime_activite_taux", "ass_aspa_taux",
            "rsa_nb", "ass_nb", "prime_activite_nb",
            "source_url", "source_millesime"]
    df = df[[c for c in keep if c in df.columns]]
    df.to_csv(PROCESSED / "minimas_sociaux.csv", index=False)
    print(f"  ✓ minimas_sociaux.csv — {len(df)} lignes")
    cov_rsa = df['rsa_taux'].notna().sum() if 'rsa_taux' in df.columns else 0
    print(f"    RSA taux couverture : {cov_rsa}/{len(df)} lignes")


def _write_minimas_placeholder(dep_ref: pd.DataFrame) -> None:
    """Écrit un CSV placeholder avec NaN pour les minimas sociaux."""
    ph = dep_ref[["dep_code"]].copy() if not dep_ref.empty else pd.DataFrame(columns=["dep_code"])
    ph["annee"] = 2021
    for col in ("rsa_taux", "prime_activite_taux", "ass_aspa_taux"):
        ph[col] = float("nan")
    ph["source_url"] = (
        "https://data.drees.solidarites-sante.gouv.fr/explore/dataset/"
        "donnees-mensuelles-sur-les-prestations-de-solidarite/"
    )
    ph["source_millesime"] = "N/A"
    ph.to_csv(PROCESSED / "minimas_sociaux.csv", index=False)
    print("  ℹ️  minimas_sociaux.csv créé avec NaN")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("  02_clean.py — Nettoyage des sources de données")
    print("=" * 60)

    clean_surendettement()
    clean_filosofi()
    clean_chomage()
    clean_rp()
    clean_minimas()

    print("\n" + "=" * 60)
    print("  Fichiers produits dans data/processed/ :")
    for fname in [
        "surendettement.csv", "filosofi.csv", "filosofi_gini.csv",
        "chomage.csv", "rp_pop_ref.csv", "rp_infracommunal.csv", "minimas_sociaux.csv",
    ]:
        path = PROCESSED / fname
        if path.exists():
            nb = sum(1 for _ in open(path)) - 1
            print(f"  ✓ {fname} ({nb} lignes)")
        else:
            print(f"  ✗ {fname} non généré")
    print()


if __name__ == "__main__":
    main()
